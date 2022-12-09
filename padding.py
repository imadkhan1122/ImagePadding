from PIL import Image
from openpyxl import load_workbook
import os
import urllib.request as urllib
import shutil
from tqdm import tqdm


def read_excel(path):
    wb = load_workbook(path)
    ws = wb['Sheet1']
    
    column = ws['A']
    column_A = [column[x].value for x in range(len(column))]
    
    column = ws['B']
    column_B = [column[x].value for x in range(len(column))]
    
    LST = []
    for i in range(len(column_A)):
        if i != 0:
            LST.append((column_A[i], column_B[i]))
    return LST
        
def download_images(path, LST):
    # split image name into two parts
    for i in tqdm(LST):
        ImageURL = i[0]
        EXT = ImageURL.split('.')[-1]
        ImageName = i[1]+'.'+EXT
        ImagePth = path+ImageName
        # download image file to images folder
        if not os.path.exists(ImagePth):
            try:
                request = urllib.Request(ImageURL, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.urlopen(request) as response, open(ImagePth, 'wb') as out_file:
                    shutil.copyfileobj(response, out_file)
                    print('image downloaded ')
            except:
                print('Image not downloaded')
        else:
            pass
    return str('Download Completed')
            
def resize_images(Orig_pth, padd_pth, excel_path):
    LST = read_excel(excel_path)
    if not os.path.exists(Orig_pth):
        os.mkdir(Orig_pth)
    check = os.listdir(Orig_pth)
    for l in LST:
        ext = l[0].split('.')[-1]
        name = l[1]+ '.'+ ext
        if name not in check:
            download_images(Orig_pth, l)
        else:
            break
    if not os.path.exists(padd_pth):
        os.mkdir(padd_pth)

    for img in tqdm(os.listdir(Orig_pth)):
        pth = Orig_pth+img
        ext = pth.split('.')[-1].lower()
        name = img.split('.')[0]
        onesideW = 0
        onesideH = 0
        try:
            im = Image.open(pth)
            width, height = im.size  
            try:   
                if width < 500 or height < 500: 
                    if width < 500:
                        paddingWidth = 500 - width
                        onesideW = round(paddingWidth/2)+1
                    if height < 500:
                        paddingHeight = 500 - height
                        onesideH = round(paddingHeight/2)+1
                    elif width > 500 and width < 2800:
                        onesideW = width    
                    elif height > 500 and height < 2800:
                        onesideH = height
                    elif width > 2800:
                        onesideW = 2800 - width
                    elif height > 2800:
                        onesideH = 2800 - height
                elif width > 2800 and height > 2800:
                    onesideW = 2800 - width
                    onesideH = 2800 - height
                elif width < 2800 and width > 500 and height < 2800 and height > 500:
                    onesideW = 0
                    onesideH = 0
            except:
                pass     
            new_width = width + onesideW + onesideW
            new_height = height + onesideH + onesideH
            result = Image.new(im.mode, (new_width, new_height), (255, 255, 255))
            result.paste(im, (onesideW, onesideH))
            if 'jpg' in ext:
                result_name = padd_pth + name+'.jpg'
                result.save(result_name)
            elif 'png' in ext:
                result_name = padd_pth + name+'.png'
                result.save(result_name)
            elif 'JPG' in ext:
                result_name = padd_pth + name+'.webp'
                result.save(result_name)
        except:
            pass
    return str('Padding Completed')
        
